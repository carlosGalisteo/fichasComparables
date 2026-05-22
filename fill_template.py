import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname
import shutil

FONT = "Segoe UI"
FS = 6
C_RED   = "C00000"
C_BLACK = "000000"
C_BLUE  = "0563C1"

MUNICIPIOS = [
    "Adeje","Arafo","Arico","Arona","Buenavista del Norte","Candelaria",
    "El Rosario","El Sauzal","El Tanque","Fasnia","Garachico","Granadilla",
    "Guía de Isora","Güímar","Icod","La Guancha","La Matanza de Acentejo",
    "La Orotava","La Victoria de Acentejo","Los Realejos","Los Silos",
    "Puerto de la Cruz","San Cristóbal de La Laguna","San Juan de la Rambla",
    "San Miguel de Abona","Santa Cruz de Tenerife","Santa Úrsula",
    "Santiago del Teide","Tacoronte","Tegueste","Vilaflor"
]

CPS_POR_MUNICIPIO = {
    "Adeje":                    ["38615","38660","38670","38677","38678","38679"],
    "Arafo":                    ["38509","38550"],
    "Arico":                    ["38580","38588","38589","38592","38593"],
    "Arona":                    ["38626","38627","38630","38631","38632","38640","38649","38650","38652"],
    "Buenavista del Norte":     ["38480","38489"],
    "Candelaria":               ["38509","38510","38520","38530","38540"],
    "El Rosario":               ["38109","38190","38290"],
    "El Sauzal":                ["38359","38360"],
    "El Tanque":                ["38435"],
    "Fasnia":                   ["38570","38579"],
    "Garachico":                ["38450","38458","38459","38460"],
    "Granadilla":               ["38594","38595","38600","38610","38611","38612","38616","38617","38618","38619"],
    "Guía de Isora":            ["38680","38683","38685","38686","38687","38688","38689"],
    "Güímar":                   ["38500","38508","38560","38590","38591"],
    "Icod":                     ["38430","38434","38438","38439"],
    "La Guancha":               ["38440","38441","38449"],
    "La Matanza de Acentejo":   ["38370","38379"],
    "La Orotava":               ["38300","38310","38311","38312","38313","38314","38315"],
    "La Victoria de Acentejo":  ["38380","38389"],
    "Los Realejos":             ["38410","38412","38413","38414","38415","38416","38417","38418","38419"],
    "Los Silos":                ["38460","38470","38479"],
    "Puerto de la Cruz":        ["38400"],
    "San Cristóbal de La Laguna": ["38108","38201","38202","38203","38204","38205","38206","38207",
                                   "38208","38240","38250","38260","38270","38291","38293","38294",
                                   "38296","38297","38320","38329","38330"],
    "San Juan de la Rambla":    ["38420","38428","38429"],
    "San Miguel de Abona":      ["38620","38628","38629","38639"],
    "Santa Cruz de Tenerife":   ["38001","38002","38003","38004","38005","38006","38007","38008",
                                  "38009","38010","38107","38110","38111","38120","38129","38130",
                                  "38139","38140","38150","38160","38170","38180","38294"],
    "Santa Úrsula":             ["38390","38398","38399"],
    "Santiago del Teide":       ["38684","38690"],
    "Tacoronte":                ["38340","38350","38355","38356","38357","38358"],
    "Tegueste":                 ["38280","38292"],
    "Vilaflor":                 ["38613","38614","38615"],
}

# Campos numéricos obligatorios que nunca deben ser "DESCONOCIDO" como número
CAMPOS_NUMERICOS = {"B3", "D3", "D5", "F5", "H5", "B6", "D6", "F6", "B9", "F9"}

# Portales conocidos que NUNCA son el comercializador
PORTALES = {"idealista", "fotocasa", "habitaclia", "pisos.com", "milanuncios",
            "yaencontré", "yaencontre", "inmofinder", "nuroa", "buscoinmueble"}


def _validate_campos(campos, flags):
    """
    Validaciones defensivas sobre el dict de campos antes de escribir.
    Lanza ValueError con mensaje descriptivo si detecta inconsistencias.
    Corrige automáticamente lo que puede; marca DESCONOCIDO lo que no puede.
    """
    errors = []
    desc = flags.get("campos_desconocidos", [])

    # 1. Precio: nunca puede ser 0 ni negativo
    precio = campos.get("B9")
    if precio is not None and precio != "DESCONOCIDO":
        try:
            p = float(precio)
            if p <= 0:
                errors.append("B9 (precio) es 0 o negativo")
        except (TypeError, ValueError):
            errors.append(f"B9 (precio) no es numérico: {precio!r}")

    # 2. Sup. construida: nunca puede ser 0 ni negativa
    sc = campos.get("B6")
    if sc is not None and sc != "DESCONOCIDO":
        try:
            s = float(sc)
            if s <= 0:
                errors.append("B6 (sup. construida) es 0 o negativa")
        except (TypeError, ValueError):
            errors.append(f"B6 (sup. construida) no es numérico: {sc!r}")

    # 3. Parcela: si existe debe ser > 0
    parcela = campos.get("H5")
    if parcela is not None and parcela != "DESCONOCIDO":
        try:
            p = float(parcela)
            if p <= 0:
                errors.append("H5 (parcela) es 0 o negativa")
        except (TypeError, ValueError):
            errors.append(f"H5 (parcela) no es numérico: {parcela!r}")

    # 4. Municipio válido
    municipio = campos.get("F2", "")
    if municipio and municipio != "DESCONOCIDO" and municipio not in MUNICIPIOS:
        errors.append(f"F2 (municipio) '{municipio}' no está en la lista de municipios válidos")

    # 5. CP coherente con municipio
    cp = campos.get("F3", "")
    if cp and cp != "DESCONOCIDO" and municipio and municipio != "DESCONOCIDO":
        cps_validos = CPS_POR_MUNICIPIO.get(municipio, [])
        if cps_validos and cp not in cps_validos:
            errors.append(f"F3 (CP) '{cp}' no corresponde al municipio '{municipio}'. "
                          f"CPs válidos: {cps_validos}")

    # 6. Comercializador no puede ser un portal conocido
    comercializador = campos.get("D11", "") or ""
    if comercializador.lower().strip() in PORTALES:
        errors.append(f"D11 (comercializador) contiene el nombre del portal '{comercializador}', "
                      f"no de la agencia. Debe ser el nombre de la inmobiliaria anunciante.")

    # 7. Planta de casa/chalet debe ser 0
    tipo = (campos.get("B5") or "").lower()
    planta = campos.get("D3")
    if any(t in tipo for t in ["casa", "chalet", "local", "nave", "terreno"]):
        if planta is not None and planta != "DESCONOCIDO" and planta != 0:
            errors.append(f"D3 (planta) = {planta} pero el tipo '{campos.get('B5')}' "
                          f"siempre tiene planta 0. Corrigiendo automáticamente.")
            campos["D3"] = 0  # autocorrección

    # 8. Ático nunca en planta 0
    if "ático" in tipo or "atico" in tipo:
        if planta == 0:
            errors.append("D3 (planta) = 0 para un ático, lo cual es imposible. "
                          "El ático está en la última planta (≥1). Marcar como DESCONOCIDO.")
            campos["D3"] = "DESCONOCIDO"
            if "D3" not in desc:
                desc.append("D3")
            flags["campos_desconocidos"] = desc

    # 9. Corrección coherente con fuente
    fuente = campos.get("B11", "API")
    f9 = campos.get("F9", 0.08)
    if fuente == "Particular" and f9 != 0.05:
        errors.append(f"F9 (corrección) = {f9} pero fuente = 'Particular'. "
                      f"Debería ser 0.05. Corrigiendo automáticamente.")
        campos["F9"] = 0.05
    elif fuente == "API" and f9 != 0.08:
        errors.append(f"F9 (corrección) = {f9} pero fuente = 'API'. "
                      f"Debería ser 0.08. Corrigiendo automáticamente.")
        campos["F9"] = 0.08

    # 10. Dormitorios no aplica a no-residenciales pero sí a viviendas
    tipos_no_residenciales = ["oficina", "local", "nave", "garaje", "terreno", "trastero", "edificio"]
    if any(t in tipo for t in tipos_no_residenciales):
        if campos.get("D5") is not None and campos.get("D5") != "DESCONOCIDO":
            errors.append(f"D5 (dormitorios) tiene valor para un inmueble no residencial '{tipo}'. "
                          f"Limpiando a null.")
            campos["D5"] = None

    return errors


def set_val(ws, coord, value, red=False, italic=False, bold=False, num_fmt=None, blue=False):
    c = ws[coord]
    c.value = "" if value is None else value
    color = C_RED if red else (C_BLUE if blue else C_BLACK)
    c.font = Font(name=FONT, size=FS, bold=bold, color=color, italic=italic)
    if num_fmt:
        c.number_format = num_fmt


def add_listas_sheet(wb, municipio, sheet_title):
    """Crea hoja oculta _LISTAS_N con municipios y CPs del municipio detectado."""
    listas_name = f"_LISTAS_{sheet_title.replace(' ', '_')}"
    if listas_name in wb.sheetnames:
        del wb[listas_name]

    ws_l = wb.create_sheet(listas_name)

    for i, m in enumerate(MUNICIPIOS, 1):
        ws_l.cell(row=i, column=1, value=m)

    cps = CPS_POR_MUNICIPIO.get(municipio, [])
    for i, cp in enumerate(cps, 1):
        ws_l.cell(row=i, column=2, value=cp)

    ws_l.sheet_state = "hidden"
    return listas_name, len(MUNICIPIOS), len(cps)


def add_dropdowns(ws, wb, municipio, sheet_title):
    """Añade validaciones usando referencias a hoja _LISTAS."""
    ws.data_validations.dataValidation = []

    listas_name, n_mun, n_cps = add_listas_sheet(wb, municipio, sheet_title)
    sheet_ref = quote_sheetname(listas_name)

    dv_mun = DataValidation(type="list",
        formula1=f"{sheet_ref}!$A$1:$A${n_mun}",
        allow_blank=True, showErrorMessage=False)
    dv_mun.sqref = "F2"
    ws.add_data_validation(dv_mun)

    if n_cps > 0:
        dv_cp = DataValidation(type="list",
            formula1=f"{sheet_ref}!$B$1:$B${n_cps}",
            allow_blank=True, showErrorMessage=False)
        dv_cp.sqref = "F3"
        ws.add_data_validation(dv_cp)

    dv_estado = DataValidation(type="list",
        formula1='"Muy malo,Malo,Medio,Bueno,Muy bueno"',
        allow_blank=True, showErrorMessage=False)
    dv_estado.sqref = "H6"
    ws.add_data_validation(dv_estado)

    dv_fuente = DataValidation(type="list",
        formula1='"API,Particular"',
        allow_blank=True, showErrorMessage=False)
    dv_fuente.sqref = "B11"
    ws.add_data_validation(dv_fuente)

    dv_bool = DataValidation(type="list",
        formula1='"Sí,No"',
        allow_blank=True, showErrorMessage=False)
    dv_bool.sqref = "B7 D7 F7 H7 B8"
    ws.add_data_validation(dv_bool)


def fill_comparable(template_path, out_path, num, campos, flags):
    # --- Validación defensiva antes de escribir nada ---
    validation_errors = _validate_campos(campos, flags)
    if validation_errors:
        print("⚠️  ADVERTENCIAS DE VALIDACIÓN (corregidas automáticamente donde fue posible):")
        for e in validation_errors:
            print(f"   • {e}")

    shutil.copy(template_path, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    sheet_title = f"COMPARABLE {num}"
    ws.title = sheet_title

    desc          = flags.get("campos_desconocidos", [])
    precio_anejos = flags.get("precio_con_anejos", False)
    municipio     = campos.get("F2", "")

    def rojo(k):
        return k in desc or (k == "B9" and precio_anejos)

    ws["A1"].value = f"COMPARABLE {num}"

    set_val(ws, "B2",  campos.get("B2"),       red=rojo("B2"))
    set_val(ws, "F2",  municipio,               red=rojo("F2"))
    set_val(ws, "B3",  campos.get("B3"),        red=rojo("B3"))
    set_val(ws, "D3",  campos.get("D3"),        red=rojo("D3"))
    set_val(ws, "F3",  campos.get("F3"),        red=rojo("F3"))
    set_val(ws, "H3",  campos.get("H3"))
    set_val(ws, "H2",  "Santa Cruz de Tenerife")
    set_val(ws, "B4",  campos.get("B4", ""))

    set_val(ws, "B5",  campos.get("B5"),        red=rojo("B5"))
    set_val(ws, "D5",  campos.get("D5"),        red=rojo("D5"))
    set_val(ws, "F5",  campos.get("F5"),        red=rojo("F5"))
    set_val(ws, "H5",  campos.get("H5"),        red=rojo("H5"))
    set_val(ws, "B6",  campos.get("B6"),        red=rojo("B6"))
    set_val(ws, "D6",  campos.get("D6"),        red=rojo("D6"))
    set_val(ws, "F6",  campos.get("F6"),        red=rojo("F6"))
    set_val(ws, "H6",  campos.get("H6", ""))

    set_val(ws, "B7",  campos.get("B7"))
    set_val(ws, "D7",  campos.get("D7"))
    set_val(ws, "F7",  campos.get("F7"))
    set_val(ws, "H7",  campos.get("H7"))
    set_val(ws, "B8",  campos.get("B8"))

    set_val(ws, "B9",  campos.get("B9"),        red=rojo("B9"), num_fmt="#,##0 €")
    ws["D9"].value = None
    ws["F9"].value = campos.get("F9", 0.08)
    ws["F9"].number_format = "0%"

    set_val(ws, "B11", campos.get("B11"))
    set_val(ws, "D11", campos.get("D11"),       red=rojo("D11"))
    set_val(ws, "B12", campos.get("B12"),       blue=True, italic=True)

    add_dropdowns(ws, wb, municipio, sheet_title)

    wb.save(out_path)
    print(f"✅ OK: {out_path}")
    if validation_errors:
        print(f"   ({len(validation_errors)} advertencia(s) registrada(s) arriba)")
