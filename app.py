"""
Comparables Inmobiliarios — Generador de Fichas de Testigos de Mercado
Tenerife · Powered by Claude Vision API
"""

import streamlit as st
import anthropic
import base64
import json
import io
import os
import shutil
import tempfile
from copy import copy as _copy
from datetime import date
from pathlib import Path
import openpyxl

from fill_template import fill_comparable

FIELD_LABELS = {
    "B2":  "Dirección / emplazamiento",
    "B3":  "Número de vía",
    "D3":  "Planta",
    "F2":  "Municipio",
    "F3":  "Código postal",
    "B4":  "Referencia catastral",
    "B5":  "Tipo de inmueble",
    "D5":  "Dormitorios",
    "F5":  "Baños",
    "H5":  "Superficie de parcela",
    "B6":  "Superficie construida",
    "D6":  "Antigüedad",
    "F6":  "Última reforma",
    "H6":  "Estado de conservación",
    "B9":  "Precio de oferta",
    "D9":  "Precio de anejos",
    "B11": "Fuente",
    "D11": "Comercializador",
    "B12": "URL del anuncio",
    "H3":  "Fecha de aportación",
    "D8":  "Ascensor",
}

TIPOLOGIAS = {
    "vivienda":      {"label": "Vivienda",        "enabled": True},
    "local_oficina": {"label": "Local / Oficina",  "enabled": False},
    "nave":          {"label": "Nave",             "enabled": False},
    "garaje":        {"label": "Garaje",           "enabled": False},
    "terreno":       {"label": "Terreno",          "enabled": False},
}

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN DE PÁGINA
# ─────────────────────────────────────────────────────────────────────────────

try:
    from PIL import Image as _PILFavicon
    _favicon = _PILFavicon.open(Path(__file__).parent / "icono.png")
except Exception:
    _favicon = "🏠"

st.set_page_config(
    page_title="Fichas comparables · CITAE",
    page_icon=_favicon,
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# CSS PERSONALIZADO
# ─────────────────────────────────────────────────────────────────────────────

st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap');

  html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
  }

  /* ── Sidebar — fondo naranja corporativo ──────────────────────────────────── */
  [data-testid="stSidebar"] {
    background: #f0f4f8;
    border-right: 1px solid #bcccdc;
  }

  /* Títulos de comparable (h2) — encabezado de bloque, prominente */
  [data-testid="stSidebar"] h2 {
    color: #101828 !important;
    font-size: 16px !important;
    font-weight: 700 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.06em !important;
    border-bottom: 2px solid #bcccdc !important;
    padding-bottom: 7px !important;
    margin-bottom: 12px !important;
    margin-top: 4px !important;
  }

  /* Etiquetas de sección (h3) — small caps */
  [data-testid="stSidebar"] h3 {
    color: #101828 !important;
    font-size: 13px !important;
    font-weight: 700 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.07em !important;
    border-bottom: 1px solid #bcccdc !important;
    padding-bottom: 5px !important;
    margin-bottom: 10px !important;
    padding-left: 0 !important;
    border-left: none !important;
  }

  /* Labels de campos */
  [data-testid="stSidebar"] label {
    color: #101828 !important;
    font-size: 12px !important;
    font-weight: 500 !important;
  }

  /* Captions */
  [data-testid="stSidebar"] [data-testid="stCaptionContainer"] {
    color: #4b5563 !important;
  }

  /* Separadores */
  [data-testid="stSidebar"] hr {
    border-color: #bcccdc !important;
    margin: 8px 0 !important;
  }

  /* Alertas (API key ok) */
  [data-testid="stSidebar"] [data-testid="stAlert"] {
    background: #e8edf3 !important;
    border: none !important;
    border-radius: 6px !important;
  }
  [data-testid="stSidebar"] [data-testid="stAlert"] p {
    color: #101828 !important;
  }

  /* Inputs de texto y fecha — fondo blanco, texto oscuro */
  [data-testid="stSidebar"] input[type="text"],
  [data-testid="stSidebar"] input[type="password"],
  [data-testid="stSidebar"] input[type="number"] {
    background: #ffffff !important;
    color: #111827 !important;
    border: none !important;
    border-radius: 6px !important;
    font-size: 13px !important;
  }
  [data-testid="stSidebar"] input::placeholder {
    color: #9ca3af !important;
  }

  /* Selectbox — fondo blanco, texto oscuro */
  [data-testid="stSidebar"] div[data-baseweb="select"] > div:first-child {
    background: #ffffff !important;
    border: none !important;
    border-radius: 6px !important;
  }
  [data-testid="stSidebar"] div[data-baseweb="select"] > div:first-child,
  [data-testid="stSidebar"] div[data-baseweb="select"] > div:first-child * {
    color: #111827 !important;
  }

  /* File uploader — dropzone */
  [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] {
    background: #ffffff !important;
    border: 1px dashed #bcccdc !important;
    border-radius: 8px !important;
  }
  /* Botón "Browse files" dentro del dropzone */
  [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button {
    background: #df7620 !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 6px !important;
  }
  [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button:hover {
    background: #c96518 !important;
  }
  [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button p,
  [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button span {
    color: #ffffff !important;
  }
  [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button svg {
    fill: #ffffff !important;
    color: #ffffff !important;
    stroke: #ffffff !important;
  }
  [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button svg path {
    fill: #ffffff !important;
    stroke: #ffffff !important;
  }
  /* Texto de instrucción del dropzone */
  [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] small,
  [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] span {
    color: #101828 !important;
  }

  /* Tarjetas de comparables guardados (filas de st.columns) */
  [data-testid="stSidebar"] [data-testid="stHorizontalBlock"] {
    background: #ffffff !important;
    border: 1px solid #bcccdc !important;
    border-radius: 8px !important;
    padding: 6px 10px !important;
    margin-bottom: 4px !important;
  }
  [data-testid="stSidebar"] [data-testid="stHorizontalBlock"] p,
  [data-testid="stSidebar"] [data-testid="stHorizontalBlock"] small {
    color: #101828 !important;
  }

  /* Botón ✕ dentro de tarjeta — rojo discreto */
  [data-testid="stSidebar"] [data-testid="stHorizontalBlock"] .stButton > button {
    background: transparent !important;
    color: #dc2626 !important;
    border: 1px solid rgba(220,38,38,0.4) !important;
    border-radius: 4px !important;
    font-size: 12px !important;
    box-shadow: none !important;
  }
  [data-testid="stSidebar"] [data-testid="stHorizontalBlock"] .stButton > button:hover {
    background: #fef2f2 !important;
    color: #b91c1c !important;
    border-color: #dc2626 !important;
  }

  /* Botón secundario en sidebar (Limpiar todo) — destructivo */
  [data-testid="stSidebar"] .stButton > button:not([kind="primary"]) {
    background: #ffffff !important;
    color: #dc2626 !important;
    border: 1px solid rgba(220,38,38,0.35) !important;
    border-radius: 6px !important;
    font-size: 13px !important;
  }
  [data-testid="stSidebar"] .stButton > button:not([kind="primary"]):hover {
    background: #fef2f2 !important;
    border-color: #dc2626 !important;
  }

  /* Espaciado compacto entre campos */
  [data-testid="stSidebar"] .stTextInput,
  [data-testid="stSidebar"] .stSelectbox,
  [data-testid="stSidebar"] .stDateInput,
  [data-testid="stSidebar"] .stFileUploader {
    margin-bottom: 6px !important;
  }

  /* ── Cabecera principal ───────────────────────────────────────────────────── */
  .app-header {
    background: #ffffff;
    border: 1px solid #e5e7eb;
    border-radius: 10px;
    padding: 18px 28px;
    margin-bottom: 16px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 48px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.04);
  }
  .app-header h1 {
    color: #111827;
    font-size: 30px;
    font-weight: 700;
    margin: 0;
    line-height: 1.2;
  }
  .app-header p {
    color: #4b5563;
    font-size: 13px;
    margin: 4px 0 0 0;
  }

  /* ── Tarjetas del resumen del lote (zona central) ────────────────────────── */
  .lote-row {
    display: flex;
    align-items: center;
    gap: 10px;
    background: #ffffff;
    border: 1px solid #e5e7eb;
    border-radius: 8px;
    padding: 8px 14px;
    margin-bottom: 6px;
  }
  .lote-num {
    background: #dd4717;
    color: #ffffff;
    font-size: 11px;
    font-weight: 700;
    min-width: 24px;
    height: 24px;
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    flex-shrink: 0;
  }
  .lote-name { font-size: 13px; font-weight: 500; color: #111827; }
  .lote-meta { font-size: 11px; color: #6b7280; margin-top: 1px; }

  /* ── Cards de resultados ─────────────────────────────────────────────────── */
  .comparable-card {
    background: #ffffff;
    border: 1px solid #e5e7eb;
    border-radius: 10px;
    padding: 14px 18px;
    margin-bottom: 10px;
    transition: border-color 0.2s;
  }
  .comparable-card:hover { border-color: #dd4717; }
  .comparable-card.done { border-color: #10b981; background: #f0fdf4; }
  .comparable-card.error { border-color: #ef4444; background: #fef2f2; }

  /* ── Badges ──────────────────────────────────────────────────────────────── */
  .badge-unknown {
    display: inline-block;
    background: #fee2e2;
    color: #991b1b;
    font-size: 10px;
    font-weight: 600;
    padding: 2px 6px;
    border-radius: 4px;
    margin: 2px;
    font-family: 'DM Mono', monospace;
  }
  .badge-ok {
    display: inline-block;
    background: #d1fae5;
    color: #065f46;
    font-size: 10px;
    font-weight: 600;
    padding: 2px 6px;
    border-radius: 4px;
    margin: 2px;
    font-family: 'DM Mono', monospace;
  }

  /* ── Tabla de datos extraídos ────────────────────────────────────────────── */
  .data-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 8px;
    margin-top: 12px;
  }
  .data-cell {
    background: #ffffff;
    border: 1px solid #e5e7eb;
    border-radius: 6px;
    padding: 8px 10px;
  }
  .data-cell .label {
    font-size: 9px;
    color: #9ca3af;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    font-weight: 600;
    margin-bottom: 2px;
  }
  .data-cell .value {
    font-size: 13px;
    color: #111827;
    font-weight: 500;
    font-family: 'DM Mono', monospace;
  }
  .data-cell .value.unknown { color: #dc2626; }

  /* ── Notas del analista ──────────────────────────────────────────────────── */
  .analyst-notes {
    background: #fffbeb;
    border-left: 3px solid #f59e0b;
    padding: 10px 14px;
    border-radius: 0 6px 6px 0;
    font-size: 12px;
    color: #6b7280;
    margin-top: 10px;
    font-style: italic;
  }

  /* ── Botón de descarga Excel — verde ─────────────────────────────────────── */
  .stDownloadButton > button {
    background: #059669 !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
  }
  .stDownloadButton > button:hover {
    background: #047857 !important;
  }

  /* ── Botón primario global — naranja corporativo ─────────────────────────── */
  .stButton > button[kind="primary"] {
    background: #dd4717 !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    letter-spacing: 0.02em !important;
    transition: background 0.18s !important;
  }
  .stButton > button[kind="primary"]:hover {
    background: #c03d11 !important;
    box-shadow: 0 3px 10px rgba(221,71,23,0.3) !important;
  }
  .stButton > button[kind="primary"]:disabled {
    background: #f3c8b5 !important;
    color: #a05030 !important;
    box-shadow: none !important;
  }

  /* ── Botón primario en sidebar — naranja corporativo ─────────────────────── */
  [data-testid="stSidebar"] .stButton > button[kind="primary"],
  [data-testid="stSidebar"] button[data-testid="baseButton-primary"] {
    background: #df7620 !important;
    color: #ffffff !important;
    border: 1.5px solid #df7620 !important;
    font-weight: 700 !important;
    box-shadow: none !important;
    transform: none !important;
  }
  [data-testid="stSidebar"] .stButton > button[kind="primary"] p,
  [data-testid="stSidebar"] button[data-testid="baseButton-primary"] p {
    color: #ffffff !important;
  }
  [data-testid="stSidebar"] .stButton > button[kind="primary"]:hover,
  [data-testid="stSidebar"] button[data-testid="baseButton-primary"]:hover {
    background: #c96518 !important;
    border-color: #c96518 !important;
    color: #ffffff !important;
    box-shadow: none !important;
    transform: none !important;
  }
  [data-testid="stSidebar"] .stButton > button[kind="primary"]:disabled,
  [data-testid="stSidebar"] button[data-testid="baseButton-primary"]:disabled {
    background: rgba(223,118,32,0.35) !important;
    border-color: transparent !important;
    color: rgba(255,255,255,0.6) !important;
    box-shadow: none !important;
  }

  /* ── Botonera de tipología ─────────────────────────────────────────────────── */
  /* Geometría base: todos los botones del panel comparten estos valores exactos. */
  .main [data-testid="stHorizontalBlock"] [data-testid="column"]:last-child button[kind] {
    width: 100% !important;
    padding: 7px 10px !important;
    font-size: 11px !important;
    font-weight: 600 !important;
    line-height: 1.3 !important;
    letter-spacing: 0.07em !important;
    text-align: center !important;
    border-radius: 12px !important;
    border-width: 1.5px !important;
    border-style: solid !important;
    box-sizing: border-box !important;
    box-shadow: none !important;
    transition: none !important;
    transform: none !important;
    outline: none !important;
  }
  /* No seleccionado (secondary) */
  .main [data-testid="stHorizontalBlock"] [data-testid="column"]:last-child button[kind="secondary"] {
    background: #ffffff !important;
    color: #111827 !important;
    border-color: #dd4717 !important;
  }
  .main [data-testid="stHorizontalBlock"] [data-testid="column"]:last-child button[kind="secondary"]:hover {
    background: #fff4f1 !important;
    border-color: #c03d11 !important;
    color: #c03d11 !important;
    box-shadow: none !important;
    transform: none !important;
  }
  .main [data-testid="stHorizontalBlock"] [data-testid="column"]:last-child button[kind="secondary"]:active,
  .main [data-testid="stHorizontalBlock"] [data-testid="column"]:last-child button[kind="secondary"]:focus {
    background: #fff4f1 !important;
    border-color: #c03d11 !important;
    color: #c03d11 !important;
    box-shadow: none !important;
    transform: none !important;
  }
  /* No seleccionado deshabilitado (bloqueado) */
  .main [data-testid="stHorizontalBlock"] [data-testid="column"]:last-child button[kind="secondary"]:disabled {
    background: #f9fafb !important;
    color: #9ca3af !important;
    border-color: #e5e7eb !important;
    cursor: not-allowed !important;
  }
  /* Seleccionado (primary) — no cambia en ningún estado */
  .main [data-testid="stHorizontalBlock"] [data-testid="column"]:last-child button[kind="primary"],
  .main [data-testid="stHorizontalBlock"] [data-testid="column"]:last-child button[kind="primary"]:hover,
  .main [data-testid="stHorizontalBlock"] [data-testid="column"]:last-child button[kind="primary"]:active,
  .main [data-testid="stHorizontalBlock"] [data-testid="column"]:last-child button[kind="primary"]:focus {
    background: #dd4717 !important;
    color: #ffffff !important;
    border-color: #dd4717 !important;
    font-weight: 700 !important;
    cursor: default !important;
    box-shadow: none !important;
    transform: none !important;
  }
  /* Separador izquierdo del panel de tipología */
  .main [data-testid="stHorizontalBlock"] [data-testid="column"]:last-child {
    border-left: 1px solid #e5e7eb;
    padding-left: 1.25rem;
  }

  /* ── Ocultar elementos de Streamlit ──────────────────────────────────────── */
  /* header NO se oculta: contiene el toggle del sidebar al plegarlo */
  #MainMenu { visibility: hidden; }
  footer { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# CARGA DEL SYSTEM PROMPT
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_data
def load_system_prompt():
    sp_path = Path(__file__).parent / "SYSTEM_PROMPT_COMPARABLES.txt"
    if sp_path.exists():
        return sp_path.read_text(encoding="utf-8")
    return ""

SYSTEM_PROMPT = load_system_prompt()

# ─────────────────────────────────────────────────────────────────────────────
# FUNCIONES CORE
# ─────────────────────────────────────────────────────────────────────────────

def image_to_base64(uploaded_file) -> tuple[str, str]:
    """Convierte un archivo subido a base64 + media_type."""
    data = uploaded_file.read()
    b64 = base64.standard_b64encode(data).decode("utf-8")
    mt = uploaded_file.type or "image/png"
    return b64, mt


def bytes_to_b64(data: bytes) -> str:
    return base64.standard_b64encode(data).decode("utf-8")


_MAX_PART_BYTES = 5 * 1024 * 1024  # 5 MB — límite de Claude Vision por imagen


def _make_image_part(img, label: str) -> dict:
    """
    Convierte un PIL Image a un ImagePart JPEG con validación.
    Recomprime progresivamente (85→80→75) si supera 5 MB.
    Devuelve dict: bytes, media_type, width, height, size_bytes, label, ok, error.
    """
    if img.mode not in ("RGB", "L"):
        img = img.convert("RGB")
    jpeg_bytes = b""
    for q in (85, 80, 75):
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=q, optimize=True)
        jpeg_bytes = buf.getvalue()
        if len(jpeg_bytes) <= _MAX_PART_BYTES:
            break
    w, h = img.size
    size = len(jpeg_bytes)
    errors = []
    if w > 7600:
        errors.append(f"ancho {w} px > 7600")
    if h > 7600:
        errors.append(f"alto {h} px > 7600")
    if size > _MAX_PART_BYTES:
        errors.append(f"peso {size / 1048576:.2f} MB > 5 MB")
    return {
        "bytes":      jpeg_bytes,
        "media_type": "image/jpeg",
        "width":      w,
        "height":     h,
        "size_bytes": size,
        "label":      label,
        "ok":         len(errors) == 0,
        "error":      "; ".join(errors),
    }


def resize_to_fit(img_bytes: bytes, label: str = "", max_dim: int = 7600) -> dict:
    """
    Redimensiona proporcionalmente si alguna dimensión supera max_dim.
    Siempre devuelve ImagePart JPEG (aunque no redimensione).
    """
    from PIL import Image as _PIL
    img = _PIL.open(io.BytesIO(img_bytes))
    w, h = img.size
    if max(w, h) > max_dim:
        scale = max_dim / max(w, h)
        img = img.resize((int(w * scale), int(h * scale)), _PIL.LANCZOS)
    return _make_image_part(img, label)


def split_or_resize(
    img_bytes: bytes,
    max_dim: int = 7600,
    resize_threshold: int = 11000,
    overlap: int = 250,
) -> list[dict]:
    """
    Devuelve lista de ImagePart JPEG:
    - [original→jpeg]  si max(w,h) <= max_dim
    - [resized→jpeg]   si max(w,h) <= resize_threshold (reducción ≤31%)
    - [banda1, ...]    si max(w,h) > resize_threshold (fragmentación con solape)
    Nunca descarta ninguna parte de la imagen.
    """
    from PIL import Image as _PIL
    img = _PIL.open(io.BytesIO(img_bytes))
    w, h = img.size
    longest = max(w, h)

    if longest <= max_dim:
        return [_make_image_part(img, "IMAGEN COMPLETA (referencia)")]

    if longest <= resize_threshold:
        scale = max_dim / longest
        img = img.resize((int(w * scale), int(h * scale)), _PIL.LANCZOS)
        return [_make_image_part(img, "IMAGEN COMPLETA (referencia, redimensionada)")]

    step = max_dim - overlap
    parts = []
    if h >= w:
        tops = []
        top = 0
        while top < h:
            bottom = min(top + max_dim, h)
            tops.append((top, bottom))
            if bottom == h:
                break
            top += step
        n = len(tops)
        for i, (top, bottom) in enumerate(tops):
            band = img.crop((0, top, w, bottom))
            parts.append(_make_image_part(
                band,
                f"Imagen completa del anuncio — fragmento {i + 1} de {n}",
            ))
    else:
        lefts = []
        left = 0
        while left < w:
            right = min(left + max_dim, w)
            lefts.append((left, right))
            if right == w:
                break
            left += step
        n = len(lefts)
        for i, (left, right) in enumerate(lefts):
            band = img.crop((left, 0, right, h))
            parts.append(_make_image_part(
                band,
                f"Imagen completa del anuncio — fragmento {i + 1} de {n}",
            ))
    return parts


def crop_zones(img_bytes: bytes) -> dict[str, tuple[bytes, str]]:
    """
    Bloque 0: recorta la imagen en zonas estándar usando PIL.
    Devuelve dict {nombre: (bytes_png, media_type)}.
    """
    from PIL import Image as PILImage
    import io as _io

    img = PILImage.open(_io.BytesIO(img_bytes))
    w, h = img.size

    zonas = {
        "ZONA_FICHA":     img.crop((0, int(h*0.03), w, int(h*0.38))),
        "ZONA_PRECIO":    img.crop((0, int(h*0.55), w, int(h*0.75))),
        "ZONA_CABECERA":  img.crop((0, 0,            w, int(h*0.06))),
        "ZONA_UBICACION": img.crop((0, int(h*0.80),  w, int(h*0.92))),
        "ZONA_ANUNCIANTE":img.crop((0, int(h*0.10),  w, int(h*0.22))),
    }

    result = {}
    for nombre, zona in zonas.items():
        buf = _io.BytesIO()
        zona.save(buf, format="PNG")
        result[nombre] = (buf.getvalue(), "image/png")
    return result


_ZONA_LABELS = {
    "ZONA_CABECERA":   "ZONA_CABECERA — dirección, título, precio destacado",
    "ZONA_FICHA":      "ZONA_FICHA — características: sup. construida, dormitorios, baños, dotaciones, año construcción, comercializador",
    "ZONA_PRECIO":     "ZONA_PRECIO — precio exacto, precio/m², anejos incluidos",
    "ZONA_UBICACION":  "ZONA_UBICACION — municipio, barrio, zona, CP",
    "ZONA_ANUNCIANTE": "ZONA_ANUNCIANTE — nombre de la agencia / inmobiliaria anunciante",
}


def prepare_image_parts(img_bytes: bytes) -> list[dict]:
    """
    Prepara todas las partes ImagePart JPEG para un comparable.
    No llama a Claude ni a Streamlit.
    Reutilizable para validación local y para el análisis real.
    """
    parts = []
    for nombre, (zona_bytes, _) in crop_zones(img_bytes).items():
        parts.append(resize_to_fit(zona_bytes, label=_ZONA_LABELS.get(nombre, nombre)))
    for fpart in split_or_resize(img_bytes):
        parts.append(fpart)
    return parts


def extract_comparable_from_image(
    client: anthropic.Anthropic,
    img_b64: str,
    img_media_type: str,
    num: int,
    url: str,
    fecha: str,
    ref_catastral: str = "",
) -> dict:
    """
    Llama a Claude con visión aplicando el preprocesado del Bloque 0:
    recorta la imagen en zonas y las envía junto con la imagen completa
    para maximizar la precisión de lectura.
    """
    # Preparar partes JPEG (zonas + imagen completa normalizada)
    _all_parts = prepare_image_parts(base64.b64decode(img_b64))

    # Construir el mensaje multi-imagen
    content_parts = []
    for _p in _all_parts:
        content_parts.append({"type": "text", "text": f"--- {_p['label']} ---"})
        content_parts.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": _p["media_type"],
                "data": bytes_to_b64(_p["bytes"]),
            },
        })

    # ── Diagnóstico local visible en Streamlit ────────────────────────────────
    with st.expander(f"🔎 Diagnóstico imágenes — comparable {num}", expanded=False):
        _diag_md = (
            "| Etiqueta | Dimensiones | Peso | Tipo | Estado |\n"
            "|---|---|---|---|---|\n"
        )
        for _p in _all_parts:
            _st = "✅ OK" if _p["ok"] else f"❌ {_p['error']}"
            _diag_md += (
                f"| {_p['label'][:45]} | {_p['width']}×{_p['height']} px"
                f" | {_p['size_bytes'] / 1048576:.2f} MB"
                f" | {_p['media_type']} | {_st} |\n"
            )
        st.markdown(_diag_md)

    # ── Validación local — no llamar a Claude si alguna parte no cumple ───────
    _bad_parts = [_p for _p in _all_parts if not _p["ok"]]
    if _bad_parts:
        _detail = "\n".join(
            f"  • {_p['label']}: {_p['width']}×{_p['height']} px, "
            f"{_p['size_bytes'] / 1048576:.2f} MB, {_p['media_type']} — {_p['error']}"
            for _p in _bad_parts
        )
        raise ValueError(
            f"Comparable {num}: partes no válidas para Claude Vision "
            f"(no se ha realizado la llamada a la API):\n{_detail}"
        )

    user_msg = f"""Se te envían 5 recortes de zonas específicas del anuncio (ZONA_CABECERA, ZONA_FICHA,
ZONA_PRECIO, ZONA_UBICACION, ZONA_ANUNCIANTE) seguidos de la imagen completa como referencia.

Lee PRIMERO cada zona recortada con atención — el texto es más legible en los recortes.
Usa la imagen completa solo para confirmar datos dudosos.

DATOS APORTADOS POR EL USUARIO:
- Número de comparable: {num}
- URL del anuncio: {url if url else "No aportada"}
- Fecha de aportación: {fecha}
- Referencia catastral: {ref_catastral if ref_catastral else ""}

INSTRUCCIONES CRÍTICAS:
1. El COMERCIALIZADOR es el nombre de la inmobiliaria o agencia anunciante visible en
   ZONA_ANUNCIANTE o ZONA_FICHA. NUNCA es el nombre del portal (Idealista, Fotocasa, etc.).
   Si aparece "RMA", "Engel & Völkers", "Re/Max" u otro nombre de agencia → úsalo.
2. Lee el precio con máxima atención en ZONA_PRECIO.
3. Lee superficies y dormitorios en ZONA_FICHA.
4. Si algún valor es ambiguo → DESCONOCIDO.

Devuelve ÚNICAMENTE el objeto JSON válido, sin texto adicional, sin marcadores de código."""

    content_parts.append({"type": "text", "text": user_msg})

    response = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=2000,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": content_parts}],
    )

    raw = response.content[0].text.strip()
    # Limpiar backticks residuales
    raw = raw.replace("```json", "").replace("```", "").strip()
    # Extraer el primer objeto JSON válido aunque haya texto alrededor
    start = raw.find("{")
    end   = raw.rfind("}") + 1
    if start == -1 or end == 0:
        raise json.JSONDecodeError("No se encontró objeto JSON en la respuesta", raw, 0)
    return json.loads(raw[start:end])


def build_xlsx_multisheet(comparables_data: list[dict]) -> bytes:
    """
    Genera un único .xlsx con una hoja por comparable.
    Usa fill_comparable() de fill_template.py sobre la plantilla base.
    """
    template_path = Path(__file__).parent / "COMPARABLE_FICHA_v4.xlsx"

    with tempfile.TemporaryDirectory() as tmpdir:
        # Generar un xlsx por comparable
        xlsx_paths = []
        for item in comparables_data:
            num    = item["comparable"]
            campos = item["campos"]
            flags  = item.get("flags", {})
            out_p  = Path(tmpdir) / f"comp_{num}.xlsx"
            fill_comparable(str(template_path), str(out_p), num, campos, flags)
            xlsx_paths.append(out_p)

        if len(xlsx_paths) == 1:
            return xlsx_paths[0].read_bytes()

        # Combinar todas las hojas en un único workbook
        wb_final = openpyxl.Workbook()
        wb_final.remove(wb_final.active)  # quitar hoja vacía inicial

        for p in xlsx_paths:
            wb_src = openpyxl.load_workbook(p)
            for shname in wb_src.sheetnames:
                ws_src = wb_src[shname]
                ws_dst = wb_final.create_sheet(title=shname)
                # Preservar estado oculto/visible (e.g. hojas _LISTAS_*)
                ws_dst.sheet_state = ws_src.sheet_state
                for row in ws_src.iter_rows():
                    for cell in row:
                        new_cell = ws_dst.cell(
                            row=cell.row, column=cell.column, value=cell.value
                        )
                        if cell.has_style:
                            new_cell.font      = cell.font.copy()
                            new_cell.fill      = cell.fill.copy()
                            new_cell.border    = cell.border.copy()
                            new_cell.alignment = cell.alignment.copy()
                            new_cell.number_format = cell.number_format
                # Copiar anchos de columna y altos de fila
                for col_dim in ws_src.column_dimensions.values():
                    ws_dst.column_dimensions[col_dim.index].width = col_dim.width
                for row_dim in ws_src.row_dimensions.values():
                    ws_dst.row_dimensions[row_dim.index].height = row_dim.height
                # Copiar celdas combinadas
                for merge_range in ws_src.merged_cells.ranges:
                    ws_dst.merge_cells(str(merge_range))
                # Copiar validaciones de datos (desplegables)
                for dv in ws_src.data_validations.dataValidation:
                    ws_dst.add_data_validation(_copy(dv))

        buf = io.BytesIO()
        wb_final.save(buf)
        return buf.getvalue()


def render_extracted_data(item: dict):
    """Muestra un resumen visual de los datos extraídos de un comparable."""
    campos = item.get("campos", {})
    flags  = item.get("flags", {})
    desc   = flags.get("campos_desconocidos", [])
    notas  = flags.get("notas", "")

    def v(key, default="—"):
        val = campos.get(key)
        if val is None or val == "":
            return default
        return str(val)

    def cls(key):
        return "unknown" if key in desc or v(key) == "DESCONOCIDO" else ""

    precio_raw = campos.get("B9")
    if isinstance(precio_raw, (int, float)):
        precio_fmt = f"{precio_raw:,.0f} €"
    else:
        precio_fmt = str(precio_raw) if precio_raw else "—"

    html = f"""
    <div class="data-grid">
      <div class="data-cell">
        <div class="label">Dirección</div>
        <div class="value {cls('B2')}">{v('B2')} {v('B3', '')}</div>
      </div>
      <div class="data-cell">
        <div class="label">Municipio</div>
        <div class="value {cls('F2')}">{v('F2')}</div>
      </div>
      <div class="data-cell">
        <div class="label">C.P.</div>
        <div class="value {cls('F3')}">{v('F3')}</div>
      </div>
      <div class="data-cell">
        <div class="label">Tipo</div>
        <div class="value {cls('B5')}">{v('B5')}</div>
      </div>
      <div class="data-cell">
        <div class="label">Estado</div>
        <div class="value">{v('H6')}</div>
      </div>
      <div class="data-cell">
        <div class="label">Planta</div>
        <div class="value {cls('D3')}">{v('D3')}</div>
      </div>
      <div class="data-cell">
        <div class="label">Dormitorios</div>
        <div class="value {cls('D5')}">{v('D5')}</div>
      </div>
      <div class="data-cell">
        <div class="label">Baños</div>
        <div class="value {cls('F5')}">{v('F5')}</div>
      </div>
      <div class="data-cell">
        <div class="label">Sup. construida</div>
        <div class="value {cls('B6')}">{v('B6')} {'m²' if v('B6') not in ('—','DESCONOCIDO') else ''}</div>
      </div>
      <div class="data-cell">
        <div class="label">Antigüedad</div>
        <div class="value {cls('D6')}">{v('D6')} {'años' if v('D6') not in ('—','DESCONOCIDO') else ''}</div>
      </div>
      <div class="data-cell">
        <div class="label">Precio oferta</div>
        <div class="value {cls('B9')}">{precio_fmt}</div>
      </div>
      <div class="data-cell">
        <div class="label">Fuente</div>
        <div class="value">{v('B11')}</div>
      </div>
      <div class="data-cell">
        <div class="label">Comercializador</div>
        <div class="value {cls('D11')}">{v('D11')}</div>
      </div>
    </div>
    """

    if desc:
        badges = "".join(
            f'<span class="badge-unknown">{FIELD_LABELS.get(c, f"Campo no identificado ({c})")}</span>'
            for c in desc
        )
        html += f'<div style="margin-top:10px;"><strong style="font-size:11px;color:#991b1b;">Campos desconocidos:</strong> {badges}</div>'

    if flags.get("precio_con_anejos"):
        html += '<div style="margin-top:6px;"><span class="badge-unknown">⚠ Precio incluye anejos — revisar precio de anejos</span></div>'

    if notas:
        html += f'<div class="analyst-notes">💬 {notas}</div>'

    st.markdown(html, unsafe_allow_html=True)



# ─────────────────────────────────────────────────────────────────────────────
# HELPERS DE ESTADO
# ─────────────────────────────────────────────────────────────────────────────

def get_review_items(item: dict) -> list[str]:
    """Devuelve lista de incidencias legibles para el usuario final."""
    flags = item.get("flags", {})
    result = []
    for code in flags.get("campos_desconocidos", []):
        result.append(FIELD_LABELS.get(code, f"Campo no identificado ({code})"))
    if flags.get("precio_con_anejos"):
        result.append(
            "Precio de oferta: incluye anejos. Revisar precio imputable "
            "a garaje, trastero u otros."
        )
    return result


def _renumber_inputs():
    """Reordena el campo num tras eliminar un comparable del lote."""
    for i, item in enumerate(st.session_state.inputs):
        item["num"] = i + 1


# ─────────────────────────────────────────────────────────────────────────────
# ESTADO DE SESIÓN
# ─────────────────────────────────────────────────────────────────────────────

if "inputs" not in st.session_state:
    st.session_state.inputs = []       # datos introducidos por el usuario

if "resultados" not in st.session_state:
    st.session_state.resultados = []   # resultados extraídos por Claude

if "fase" not in st.session_state:
    st.session_state.fase = "entrada"  # entrada | revision | analisis | resultados

if "form_key" not in st.session_state:
    st.session_state.form_key = 0      # incrementar al guardar para resetear el formulario

if "tipologia_actual" not in st.session_state:
    st.session_state.tipologia_actual = "vivienda"

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR — CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────────────────

# Leer API key desde Streamlit Secrets antes de renderizar el sidebar,
# para que api_key quede disponible globalmente tras el bloque with st.sidebar.
try:
    _secret_key = st.secrets.get("ANTHROPIC_API_KEY", "")
except Exception:
    _secret_key = ""

with st.sidebar:
    st.markdown("### Datos del lote")

    fecha_aportacion = st.date_input(
        "Fecha de aportación",
        value=date.today(),
        format="DD/MM/YYYY",
        help="Fecha en que se aportan las imágenes (campo H3 de la ficha)",
    )
    fecha_str = fecha_aportacion.strftime("%d/%m/%Y")

    # ── Formulario de entrada comparable a comparable ──────────────────────────
    st.markdown("---")
    _next_num = len(st.session_state.inputs) + 1
    st.markdown(f"## COMPARABLE {_next_num}")

    _uploaded = st.file_uploader(
        "Imagen del anuncio",
        type=["png", "jpg", "jpeg", "webp"],
        key=f"img_upload_{st.session_state.form_key}",
    )
    _url_input = st.text_input(
        "URL del anuncio",
        placeholder="https://www.idealista.com/inmueble/...",
        key=f"url_{st.session_state.form_key}",
    )
    _ref_input = st.text_input(
        "Ref. catastral (opcional)",
        placeholder="0000001AA0000S0000AA",
        key=f"ref_{st.session_state.form_key}",
    )
    _estado_input = st.selectbox(
        "Estado de conservación",
        options=["Muy bueno", "Bueno", "Medio", "Malo", "Muy malo"],
        index=1,
        key=f"estado_{st.session_state.form_key}",
    )

    if st.button(
        "Guardar comparable",
        type="primary",
        use_container_width=True,
        disabled=(_uploaded is None),
        help="Sube una imagen para poder guardar el comparable.",
    ):
        st.session_state.inputs.append({
            "num": _next_num,
            "image_bytes": _uploaded.read(),
            "image_name": _uploaded.name,
            "image_type": _uploaded.type or "image/png",
            "url": _url_input.strip(),
            "ref": _ref_input.strip(),
            "estado": _estado_input,
        })
        st.session_state.form_key += 1
        st.rerun()

    # ── Resumen del lote en el sidebar ─────────────────────────────────────────
    if st.session_state.inputs:
        st.markdown("---")
        st.markdown(f"### Lote ({len(st.session_state.inputs)} comparable(s))")

        for _item in st.session_state.inputs:
            _col1, _col2 = st.columns([4, 1])
            with _col1:
                st.markdown(f"**{_item['num']}.** {_item['image_name']}")
                _url_ok = "✓" if _item["url"] else "—"
                st.caption(f"Estado: {_item['estado']} · URL: {_url_ok}")
            with _col2:
                if st.button("✕", key=f"del_{_item['num']}"):
                    st.session_state.inputs = [
                        x for x in st.session_state.inputs if x["num"] != _item["num"]
                    ]
                    _renumber_inputs()
                    st.rerun()

        st.markdown("")
        if st.button("🗑 Limpiar todo", use_container_width=True):
            st.session_state.inputs = []
            st.session_state.resultados = []
            st.session_state.form_key += 1
            st.rerun()

    st.markdown("---")
    with st.expander("Configuración avanzada", expanded=False):
        if _secret_key:
            api_key = _secret_key
            st.success("API Key configurada")
        else:
            api_key = st.text_input(
                "API Key de Anthropic",
                type="password",
                placeholder="sk-ant-...",
                help="Necesaria para analizar imágenes con Claude. No se almacena.",
            )

    st.markdown(
        "<div style='font-size:11px;color:#9ca3af;text-align:center;margin-top:8px;'>"
        "Fichas Comparables · CITAE arquitectura<br>"
        "Procesamiento asistido por Claude Vision</div>",
        unsafe_allow_html=True,
    )


# ─────────────────────────────────────────────────────────────────────────────
# CABECERA PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

# Logo en base64 para incrustar en el header
import base64 as _b64
from pathlib import Path as _Path
_logo_path = _Path(__file__).parent / "CITAE.png"
if _logo_path.exists():
    _logo_b64 = _b64.b64encode(_logo_path.read_bytes()).decode()
    _logo_html = f'<img src="data:image/png;base64,{_logo_b64}" style="height:56px;width:auto;">'
else:
    _logo_html = '<div style="font-size:36px;">🏠</div>'

st.markdown(f"""
<div class="app-header">
  {_logo_html}
  <div>
    <h1>Fichas de Testigos de Mercado</h1>
    <p>Extracción automática de datos de ofertas inmobiliarias</p>
  </div>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# RESUMEN DEL LOTE Y BOTÓN DE ANÁLISIS
# ─────────────────────────────────────────────────────────────────────────────

st.markdown("---")
col_main, col_tipo = st.columns([3, 1])

# ── Panel de tipología (columna derecha) ──────────────────────────────────────
with col_tipo:
    _tipo_locked = bool(st.session_state.inputs)
    st.markdown(
        "<div style='font-size:11px;font-weight:700;text-transform:uppercase;"
        "letter-spacing:0.07em;color:#6b7280;margin-bottom:10px;'>"
        "Tipología del lote</div>",
        unsafe_allow_html=True,
    )
    for _k, _tinfo in TIPOLOGIAS.items():
        _tsel = (_k == st.session_state.tipologia_actual)
        _clicked = st.button(
            _tinfo["label"].upper(),
            key=f"tipo_{_k}",
            type="primary" if _tsel else "secondary",
            disabled=_tipo_locked,
            use_container_width=True,
        )
        if _clicked and not _tsel:
            st.session_state.tipologia_actual = _k
            st.rerun()
    if _tipo_locked:
        st.caption("Para cambiar la tipología, limpia primero el lote.")
    elif not TIPOLOGIAS[st.session_state.tipologia_actual]["enabled"]:
        st.warning("Esta tipología no está disponible todavía. "
                   "El análisis quedará deshabilitado.")

# ── Lote y acciones (columna principal) ───────────────────────────────────────
with col_main:
    if st.session_state.inputs:
        _n = len(st.session_state.inputs)
        st.markdown(f"### Lote cargado — {_n} comparable(s)")

        _lote_html = ""
        for _item in st.session_state.inputs:
            _url_ok = "✓" if _item["url"] else "—"
            _ref_raw = _item.get("ref", "")
            _ref_disp = (_ref_raw[:18] + "…") if len(_ref_raw) > 18 else (_ref_raw or "—")
            _lote_html += (
                f'<div class="lote-row">'
                f'<div class="lote-num">{_item["num"]}</div>'
                f'<div>'
                f'<div class="lote-name">{_item["image_name"]}</div>'
                f'<div class="lote-meta">Estado: {_item["estado"]} &nbsp;·&nbsp; URL: {_url_ok} &nbsp;·&nbsp; Ref: {_ref_disp}</div>'
                f'</div></div>'
            )
        st.markdown(_lote_html, unsafe_allow_html=True)

        st.markdown("---")
        _can_analyze = (
            bool(api_key)
            and bool(st.session_state.inputs)
            and TIPOLOGIAS[st.session_state.tipologia_actual]["enabled"]
        )

        if st.button(
            "🔎 Validar imágenes sin llamar a Claude",
            disabled=not st.session_state.inputs,
            help="Comprueba que todas las imágenes cumplen los límites de Claude Vision sin gastar tokens.",
            use_container_width=True,
        ):
            _val_ok = True
            for _vinp in st.session_state.inputs:
                _vparts = prepare_image_parts(_vinp["image_bytes"])
                _vbad   = [_p for _p in _vparts if not _p["ok"]]
                _vicon  = "✅" if not _vbad else "❌"
                with st.expander(
                    f"{_vicon} Comparable {_vinp['num']} — {_vinp['image_name']}",
                    expanded=bool(_vbad),
                ):
                    _vmd = (
                        "| Etiqueta | Dimensiones | Peso | Tipo | Estado |\n"
                        "|---|---|---|---|---|\n"
                    )
                    for _p in _vparts:
                        _vst = "✅ OK" if _p["ok"] else f"❌ {_p['error']}"
                        _vmd += (
                            f"| {_p['label'][:45]} | {_p['width']}×{_p['height']} px"
                            f" | {_p['size_bytes'] / 1048576:.2f} MB"
                            f" | {_p['media_type']} | {_vst} |\n"
                        )
                    st.markdown(_vmd)
                if _vbad:
                    _val_ok = False
            if _val_ok:
                st.success("✅ Todas las imágenes son válidas para Claude Vision.")

        if not api_key:
            _help_btn = "Introduce una API Key de Anthropic en el panel izquierdo."
        elif not TIPOLOGIAS[st.session_state.tipologia_actual]["enabled"]:
            _help_btn = "Esta tipología no está disponible todavía."
        else:
            _help_btn = ""

        if st.button(
            f"▶ Analizar {_n} comparable(s) con Claude",
            type="primary",
            disabled=not _can_analyze,
            help=_help_btn,
        ):
            _client = anthropic.Anthropic(api_key=api_key)
            st.session_state.resultados = []
            _errors = []

            _progress = st.progress(0, text="Preparando análisis...")
            _status = st.empty()

            for _i, _inp in enumerate(st.session_state.inputs):
                _status.info(f"Analizando comparable {_inp['num']} de {_n}…")
                try:
                    _img_b64 = base64.standard_b64encode(_inp["image_bytes"]).decode("utf-8")
                    _result  = extract_comparable_from_image(
                        _client,
                        _img_b64,
                        _inp["image_type"],
                        _inp["num"],
                        _inp["url"],
                        fecha_str,
                        _inp["ref"],
                    )
                    _campos = _result.get("campos", {})
                    _flags  = _result.get("flags", {})
                    _campos["H6"] = _inp["estado"]
                    _campos["B4"] = _inp["ref"]
                    _campos["B12"] = _inp["url"]
                    _campos["H3"] = fecha_str
                    st.session_state.resultados.append({
                        "comparable": _inp["num"],
                        "campos":     _campos,
                        "flags":      _flags,
                        "img_name":   _inp["image_name"],
                    })
                except json.JSONDecodeError as _e:
                    _errors.append(f"Comparable {_inp['num']}: JSON inválido — {_e}")
                except anthropic.APIError as _e:
                    _errors.append(f"Comparable {_inp['num']}: Error de API — {_e}")
                except Exception as _e:
                    _errors.append(f"Comparable {_inp['num']}: Error inesperado — {_e}")
                _progress.progress((_i + 1) / _n)

            _progress.empty()
            _status.empty()

            if st.session_state.resultados:
                st.success(
                    f"✅ Análisis completado — {len(st.session_state.resultados)} "
                    f"comparable(s) procesado(s) correctamente."
                )
            for _err in _errors:
                st.error(_err)
            st.session_state.fase = "resultados"
    else:
        st.info("Usa el panel izquierdo para añadir comparables al lote.")


# ─────────────────────────────────────────────────────────────────────────────
# RESULTADOS
# ─────────────────────────────────────────────────────────────────────────────

if st.session_state.resultados:
    st.markdown("---")
    st.markdown(f"### 📊 Resultados — {len(st.session_state.resultados)} comparable(s)")

    for item in st.session_state.resultados:
        num       = item.get("comparable", "?")
        img_name  = item.get("img_name", "")
        campos    = item.get("campos", {})
        flags     = item.get("flags", {})
        desc      = flags.get("campos_desconocidos", [])
        tipo      = campos.get("B5", "—")
        municipio = campos.get("F2", "—")
        precio    = campos.get("B9")

        if isinstance(precio, (int, float)):
            precio_label = f"{precio:,.0f} €"
        else:
            precio_label = str(precio) if precio else "—"

        status_cls = "error" if desc else "done"
        icon = "⚠️" if desc else "✅"

        with st.expander(
            f"{icon} **COMPARABLE {num}** · {tipo} · {municipio} · {precio_label} — {img_name}",
            expanded=True,
        ):
            render_extracted_data(item)

    # ── Generación y descarga del XLSX ──────────────────────────────────────
    st.markdown("---")
    col_dl1, col_dl2, col_dl3 = st.columns([2, 2, 1])

    with col_dl1:
        nombre_archivo = st.text_input(
            "Nombre del archivo de salida",
            value=f"Comparables_{fecha_str.replace('/', '-')}",
            placeholder="Comparables_22-05-2026",
        )

    with col_dl2:
        try:
            xlsx_bytes = build_xlsx_multisheet(st.session_state.resultados)
            fname = f"{nombre_archivo}.xlsx"

            st.download_button(
                label=f"⬇ Descargar {fname}",
                data=xlsx_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"Error al generar el XLSX: {e}")

    with col_dl3:
        total_incidencias = sum(
            len(get_review_items(item))
            for item in st.session_state.resultados
        )
        st.metric("Incidencias a revisar", total_incidencias, delta=None)

    # ── JSON bruto (expander colapsado) ──────────────────────────────────────
    with st.expander("🔎 Ver JSON extraído (debug)", expanded=False):
        st.json(st.session_state.resultados)
